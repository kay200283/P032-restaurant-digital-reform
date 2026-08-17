#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dim_date backend v3 - fiscal year extend, no import/template, fixed stats, fixed WOM"""

from flask import Blueprint, render_template, request, jsonify, session, redirect, send_file
import sqlite3
import os
from datetime import datetime, timedelta, date
import tempfile
import json
import re

dim_date_bp = Blueprint('dim_date', __name__)

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'database.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def check_admin():
    if 'user' not in session:
        return redirect('/login')
    return None

@dim_date_bp.route('/dim_date')
def dim_date_page():
    chk = check_admin()
    if chk:
        return chk
    return render_template('dim_date.html', user=session.get('user'))

ALL_COLS = 'date,year,month,week_num,weekday,fiscal_year,period,holiday_type,year_month,specific_week,wage_type,solar_term,week_of_month,jp_holiday'

# WOM formula: Monday=start of week
def calc_wom(d):
    return (d.day + d.replace(day=1).weekday() - 1) // 7 + 1

WD_MAP = {0:'周一',1:'周二',2:'周三',3:'周四',4:'周五',5:'周六',6:'周日'}

@dim_date_bp.route('/api/dim_date/list', methods=['GET'])
def api_dim_date_list():
    if 'user' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    fiscal_year = request.args.get('fiscal_year')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    sort_col = request.args.get('sort', 'date')
    sort_dir = request.args.get('dir', 'asc')
    allowed = {'date','year','month','week_num','weekday','fiscal_year','period',
               'holiday_type','year_month','specific_week','wage_type','solar_term','week_of_month','jp_holiday'}
    if sort_col not in allowed:
        sort_col = 'date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    if sort_col == 'weekday':
        sort_expr = "CASE weekday WHEN '周一' THEN 1 WHEN '周二' THEN 2 WHEN '周三' THEN 3 WHEN '周四' THEN 4 WHEN '周五' THEN 5 WHEN '周六' THEN 6 WHEN '周日' THEN 7 END"
    elif sort_col == 'wage_type':
        sort_expr = "COALESCE(wage_type, 0)"
    else:
        sort_expr = sort_col
    conn = get_db()
    c = conn.cursor()
    where_parts = []
    params = []
    if year:
        where_parts.append("year = ?")
        params.append(year)
    if month:
        where_parts.append("month = ?")
        params.append(month)
    if fiscal_year:
        where_parts.append("fiscal_year = ?")
        params.append(fiscal_year)
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    c.execute(f"SELECT COUNT(*) FROM dws_dim_date{where_sql}", params)
    total = c.fetchone()[0]
    offset = (page - 1) * per_page
    c.execute(f"SELECT {ALL_COLS} FROM dws_dim_date{where_sql} ORDER BY {sort_expr} {sort_dir} LIMIT ? OFFSET ?", params + [per_page, offset])
    rows = [dict(r) for r in c.fetchall()]
    # Stats
    c.execute(f"""SELECT
        SUM(CASE WHEN holiday_type='工作日' THEN 1 ELSE 0 END) as workdays,
        SUM(CASE WHEN holiday_type LIKE '%补班%' THEN 1 ELSE 0 END) as makeup_days,
        SUM(CASE WHEN holiday_type='假日' THEN 1 ELSE 0 END) as holiday_days,
        SUM(CASE WHEN holiday_type='春节' THEN 1 ELSE 0 END) as spring,
        SUM(CASE WHEN holiday_type='国庆' THEN 1 ELSE 0 END) as national,
        SUM(CASE WHEN holiday_type='五一' THEN 1 ELSE 0 END) as labor,
        SUM(CASE WHEN holiday_type='元旦' THEN 1 ELSE 0 END) as newyear,
        SUM(CASE WHEN holiday_type='清明' THEN 1 ELSE 0 END) as qingming,
        SUM(CASE WHEN holiday_type='端午' THEN 1 ELSE 0 END) as dragon,
        SUM(CASE WHEN holiday_type='中秋' THEN 1 ELSE 0 END) as midautumn,
        SUM(CASE WHEN wage_type=2 THEN 1 ELSE 0 END) as double_pay,
        SUM(CASE WHEN wage_type=3 THEN 1 ELSE 0 END) as triple_pay
    FROM dws_dim_date{where_sql}""", params)
    row = c.fetchone()
    stats = {k: (row[k] or 0) for k in row.keys()} if row else {}
    conn.close()
    return jsonify({'rows': rows, 'total': total, 'page': page, 'per_page': per_page, 'stats': stats})

@dim_date_bp.route('/api/dim_date/years', methods=['GET'])
def api_dim_date_years():
    if 'user' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT year FROM dws_dim_date ORDER BY year")
    years = [r[0] for r in c.fetchall()]
    c.execute("SELECT DISTINCT fiscal_year FROM dws_dim_date ORDER BY fiscal_year")
    fiscal_years = [r[0] for r in c.fetchall()]
    # Compute extend options: next 5 fiscal years after current max
    max_fy = 0
    for fy in fiscal_years:
        m = re.match(r'(\d+)年期', fy)
        if m:
            max_fy = max(max_fy, int(m.group(1)))
    extend_options = [f"{i}年期" for i in range(max_fy + 1, max_fy + 6)]
    conn.close()
    return jsonify({'years': years, 'fiscal_years': fiscal_years, 'extend_options': extend_options})

@dim_date_bp.route('/api/dim_date/export', methods=['GET'])
def api_dim_date_export():
    if 'user' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    fiscal_year = request.args.get('fiscal_year')
    export_all = request.args.get('all', '0')
    conn = get_db()
    c = conn.cursor()
    where_parts = []
    params = []
    if export_all != '1':
        if year:
            where_parts.append("year = ?")
            params.append(year)
        if month:
            where_parts.append("month = ?")
            params.append(month)
        if fiscal_year:
            where_parts.append("fiscal_year = ?")
            params.append(fiscal_year)
    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    c.execute(f"SELECT {ALL_COLS} FROM dws_dim_date{where_sql} ORDER BY date", params)
    rows = c.fetchall()
    conn.close()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active
    headers = ['日期','年','月','周别','星期','财年','期','节假日','年月','具体周','三薪标注','节气','月内周序','日本祝日']
    ws.append(headers)
    wage_map = {3: '3(三薪)', 2: '2(二薪)'}
    for r in rows:
        wage_str = wage_map.get(r[10], '') if r[10] else ''
        ws.append([r[0],r[1],r[2],r[3],r[4],r[5],r[6],r[7],r[8],r[9],wage_str,r[11],r[12],r[13]])
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2D4A3E', end_color='2D4A3E', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name); tmp.close()
    filename = f"日期维度表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _sync_year_holidays(year):
    """Sync holidays for a given year. Raises ValueError if year not supported by chinese_calendar."""
    try:
        import chinese_calendar as cal
        test_date = date(year, 1, 1)
        cal.get_holiday_detail(test_date)
    except (ImportError, NotImplementedError):
        raise ValueError(f'chinese-calendar库不支持{year}年（仅支持2004~2026年），节假日和三薪数据无法自动同步')
    from collections import defaultdict
    import lunardate
    import jpholiday
    from borax.calendars.festivals2 import TermFestival
    name_map = {"New Year's Day": "元旦", "Spring Festival": "春节",
                "Tomb-sweeping Day": "清明", "Labour Day": "五一",
                "Dragon Boat Festival": "端午", "Mid-autumn Festival": "中秋",
                "National Day": "国庆"}
    holiday_days = {}
    for h in cal.Holiday:
        holiday_days[h.value] = h.days
    by_year_name = defaultdict(list)
    for dt_h, name in cal.holidays.items():
        if dt_h.year == year:
            by_year_name[(dt_h.year, name)].append(dt_h)
    triple_pay = set()
    double_pay = set()
    for (yr, name), dates in by_year_name.items():
        dates.sort()
        n = holiday_days.get(name, 1)
        for i, dt in enumerate(dates):
            if i < n:
                triple_pay.add(dt)
            else:
                double_pay.add(dt)
    makeup_days = set()
    for dt_w, name in cal.workdays.items():
        if dt_w.year == year and name is not None:
            makeup_days.add(dt_w)
    TERM_NAMES = ['小寒','大寒','立春','雨水','惊蛰','春分','清明','谷雨',
                  '立夏','小满','芒种','夏至','小暑','大暑','立秋','处暑',
                  '白露','秋分','寒露','霜降','立冬','小雪','大雪','冬至']
    term_map = {}
    for t in TERM_NAMES:
        try:
            f = TermFestival(t)
            d = f.list_days(date(year,1,1), date(year,12,31))
            if d:
                term_map[d[0].solar.strftime('%Y-%m-%d')] = t
        except:
            pass
    conn = get_db()
    c = conn.cursor()
    updated = 0
    current = date(year, 1, 1)
    end = date(year, 12, 31)
    while current <= end:
        date_str = current.strftime('%Y-%m-%d')
        on_holiday, holiday_name = cal.get_holiday_detail(current)
        if current in makeup_days:
            holiday_type = "工作日-补班"
            wage = None
        elif on_holiday:
            cn_name = name_map.get(holiday_name, "假日")
            holiday_type = cn_name
            wage = 3 if current in triple_pay else (2 if current in double_pay else None)
        else:
            holiday_type = "工作日"
            wage = None
        ld = lunardate.LunarDate.from_solar_date(current.year, current.month, current.day)
        lunar_str = f'闰{ld.month}月{ld.day}日' if ld.is_leap_month else f'{ld.month}月{ld.day}日'
        term = term_map.get(date_str, None)
        wom = calc_wom(current)
        jp_name = jpholiday.is_holiday_name(current)
        c.execute("UPDATE dws_dim_date SET holiday_type=?, wage_type=?, lunar_date=?, solar_term=?, week_of_month=?, jp_holiday=? WHERE date=?",
                  (holiday_type, wage, lunar_str, term, wom, jp_name, date_str))
        if c.rowcount > 0:
            updated += 1
        current += timedelta(days=1)
    conn.commit()
    conn.close()
    return updated

@dim_date_bp.route('/api/dim_date/sync_holidays', methods=['POST'])
def api_dim_date_sync_holidays():
    if 'user' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    data = request.get_json(silent=True) or {}
    year = data.get('year')
    if year is not None:
        year = int(year)
    if not year:
        return jsonify({'error': '请先选择年份'}), 400
    try:
        updated = _sync_year_holidays(year)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'success': True, 'year': year, 'updated': updated})

def _calc_fiscal_year_range(fy_str):
    """Parse fiscal year string like '27年期' to (start_date, end_date)."""
    m = re.match(r'(\d+)年期', fy_str)
    if not m:
        return None, None
    fy_num = int(m.group(1))
    start = date(2000 + fy_num - 1, 9, 1)
    end = date(2000 + fy_num, 8, 31)
    return start, end

@dim_date_bp.route('/api/dim_date/extend', methods=['POST'])
def api_dim_date_extend():
    if 'user' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    data = request.get_json(silent=True) or {}
    fiscal_year = data.get('fiscal_year')
    if not fiscal_year:
        return jsonify({'error': '请选择财年'}), 400
    start, end = _calc_fiscal_year_range(fiscal_year)
    if not start:
        return jsonify({'error': '财年格式错误'}), 400
    import lunardate
    import jpholiday
    from borax.calendars.festivals2 import TermFestival
    # Pre-compute solar terms for all years in the range
    TERM_NAMES = ['小寒','大寒','立春','雨水','惊蛰','春分','清明','谷雨',
                  '立夏','小满','芒种','夏至','小暑','大暑','立秋','处暑',
                  '白露','秋分','寒露','霜降','立冬','小雪','大雪','冬至']
    term_map = {}
    for y in range(start.year, end.year + 1):
        for t in TERM_NAMES:
            try:
                f = TermFestival(t)
                d = f.list_days(date(y,1,1), date(y,12,31))
                if d:
                    term_map[d[0].solar.strftime('%Y-%m-%d')] = t
            except:
                pass
    conn = get_db()
    c = conn.cursor()
    inserted = 0
    current = start
    while current <= end:
        date_str = current.strftime('%Y-%m-%d')
        c.execute("SELECT 1 FROM dws_dim_date WHERE date=?", (date_str,))
        if c.fetchone():
            current += timedelta(days=1)
            continue
        y = current.year; m = current.month
        week_num = current.isocalendar()[1]
        weekday_cn = WD_MAP[current.weekday()]
        year_month = current.strftime('%Y-%m')
        monday = current - timedelta(days=current.weekday())
        specific_week = monday.strftime('%y/%m/%d') + '周'
        if m >= 9: fy = f"{y-2000+1}年期"; period = "上期"
        elif m >= 3: fy = f"{y-2000}年期"; period = "下期"
        else: fy = f"{y-2000}年期"; period = "上期"
        ld = lunardate.LunarDate.from_solar_date(y, m, current.day)
        lunar_str = f'闰{ld.month}月{ld.day}日' if ld.is_leap_month else f'{ld.month}月{ld.day}日'
        term = term_map.get(date_str, None)
        wom = calc_wom(current)
        jp_name = jpholiday.is_holiday_name(current)
        holiday_type = '假日' if current.weekday() >= 5 else '工作日'
        c.execute("""INSERT OR IGNORE INTO dws_dim_date
            (date,year,month,week_num,weekday,fiscal_year,period,holiday_type,year_month,specific_week,wage_type,lunar_date,solar_term,week_of_month,jp_holiday)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (date_str,y,m,week_num,weekday_cn,fy,period,holiday_type,year_month,specific_week,None,lunar_str,term,wom,jp_name))
        inserted += 1
        current += timedelta(days=1)
    conn.commit()
    conn.close()
    # Try to sync holidays for each year in the range
    sync_results = {}
    for y in range(start.year, end.year + 1):
        try:
            cnt = _sync_year_holidays(y)
            sync_results[str(y)] = f'{cnt}天已同步'
        except ValueError as e:
            sync_results[str(y)] = str(e)
        except Exception as e:
            sync_results[str(y)] = f'同步失败: {str(e)}'
    return jsonify({'success': True, 'fiscal_year': fiscal_year, 'inserted': inserted, 'sync_results': sync_results})

